"""
Smart Import Service
Phase 10A - Smart Portfolio Import Engine

Purpose:
- Read portfolio data from CSV, Excel, PDF text, and image screenshots.
- Extract likely holdings:
    Symbol
    Share Name / Company Name
    Shares
    Average Price
    Current Price
- Return clean preview rows.
- Do NOT directly update portfolio from this service.

Important:
- CSV and Excel are most reliable.
- PDF text extraction requires at least one PDF text library:
    pip install pypdf
  Optional fallbacks:
    pip install PyPDF2 pdfplumber pymupdf
- Image/JPG/PNG OCR requires:
    pillow, pytesseract, Tesseract OCR installed on Windows
"""

import csv
import os
import re
from pathlib import Path


SMART_IMPORT_SERVICE_VERSION = "PSX_STATEMENT_V3_JS_INVESTPRO"

SUPPORTED_EXTENSIONS = [
    ".csv",
    ".xlsx",
    ".xlsm",
    ".pdf",
    ".png",
    ".jpg",
    ".jpeg",
    ".bmp",
    ".tif",
    ".tiff",
]


SYMBOL_HEADERS = [
    "symbol",
    "stock",
    "scrip",
    "ticker",
    "security",
    "security code",
    "company code",
]

NAME_HEADERS = [
    "share name",
    "company name",
    "stock name",
    "security name",
    "name",
    "company",
    "share",
]

SHARES_HEADERS = [
    "shares",
    "share",
    "quantity",
    "qty",
    "holding",
    "holdings",
    "units",
    "volume",
    "balance",
]

AVG_PRICE_HEADERS = [
    "avg price",
    "average price",
    "avg cost",
    "average cost",
    "cost price",
    "purchase price",
    "buy price",
    "rate",
    "avg rate",
]

CURRENT_PRICE_HEADERS = [
    "current price",
    "market price",
    "last price",
    "close price",
    "ltp",
    "price",
    "current rate",
    "market rate",
]

INVESTMENT_HEADERS = [
    "investment",
    "cost value",
    "cost",
    "book value",
    "purchase value",
    "total cost",
]

CURRENT_VALUE_HEADERS = [
    "current value",
    "market value",
    "value",
    "market worth",
    "holding value",
]


COMMON_WORDS_TO_IGNORE = {
    "symbol",
    "stock",
    "shares",
    "quantity",
    "qty",
    "price",
    "current",
    "average",
    "market",
    "value",
    "investment",
    "profit",
    "loss",
    "total",
    "portfolio",
    "statement",
    "client",
    "account",
    "date",
    "name",
    "balance",
    "security",
    "share",
    "company",
    "broker",
    "cdc",
    "nccpl",
    "js",
    "global",
    "whatsapp",
    "image",
    "generated",
    "page",
    "report",
    "from",
    "to",
    "debit",
    "credit",
    "opening",
    "closing",
    "cash",
    "ledger",
    "exchange",
    "karachi",
    "limited",
    "ltd",
    "and",
    "the",
    "for",
    "with",
    "this",
    "that",
    "your",
    "ours",
    "at",
    "am",
    "pm",
    "jan",
    "feb",
    "mar",
    "apr",
    "may",
    "jun",
    "jul",
    "aug",
    "sep",
    "oct",
    "nov",
    "dec",
    "house", "mohalla", "colony", "jhang", "pakistan", "near", "cycle", "mor", "road", "street", "city", "customer", "id", "portfolio", "registered", "dispatch", "applicable", "balance", "summary", "funds", "type", "units", "nav", "inflows", "outflows", "purchase", "redemption", "recd", "raast", "opening", "net", "charges", "tax", "load",
}


PORTFOLIO_HEADER_KEYWORDS = [
    "symbol",
    "scrip",
    "ticker",
    "stock",
    "security",
    "qty",
    "quantity",
    "shares",
    "holding",
    "avg",
    "average",
    "cost",
    "rate",
    "current",
    "market",
    "price",
    "ltp",
]

PORTFOLIO_STOP_KEYWORDS = [
    "total",
    "grand total",
    "summary",
    "cash balance",
    "ledger",
    "dividend",
    "note",
    "disclaimer",
    "page",
    "generated",
    "client detail",
    "account detail",
]

NOISE_LINE_KEYWORDS = [
    "whatsapp",
    "image",
    "generated on",
    "client",
    "account",
    "statement",
    "portfolio value",
    "total portfolio",
    "cash balance",
    "ledger",
    "disclaimer",
    "page",
    "date",
    "from",
    "to",
    "phone",
    "email",
    "address",
]


# Common PSX symbols help OCR/PDF parser avoid random words.
# Unknown symbols can still be accepted when they appear inside a detected portfolio table.
COMMON_PSX_SYMBOLS = {
    "AABS", "ABOT", "ACPL", "ADMM", "AGHA", "AGIL", "AGP", "AICL", "AIRLINK",
    "AKBL", "ALAC", "ALTN", "ANL", "APL", "ARPL", "ASC", "ATLH", "ATRL",
    "AVN", "BAFL", "BAHL", "BATA", "BIPL", "BNWM", "BOP", "BWCL", "CEPB",
    "CHCC", "CNERGY", "COLG", "CPHL", "CSAP", "DAWH", "DGKC", "DCR",
    "EFERT", "EFUG", "ENGRO", "ENGROH", "EPCL", "FABL", "FATIMA", "FCCL",
    "FCEPL", "FFBL", "FFC", "FHAM", "FML", "GAL", "GATM", "GHGL", "GLAXO",
    "GLXO", "HBL", "HCAR", "HINOON", "HMB", "HUBC", "IBFL", "ILP", "INIL",
    "ISL", "JVDC", "KAPCO", "KEL", "KOHC", "KOHE", "LOTCHEM", "LUCK",
    "MARI", "MCB", "MEBL", "MLCF", "MTL", "MUGHAL", "NATF", "NBP", "NCL",
    "NML", "OGDC", "PAEL", "PAKT", "PABC", "PIOC", "POL", "PPL", "PSO",
    "PSX", "PTC", "SAZEW", "SEARL", "SHEL", "SHFA", "SNGP", "SSGC", "SYS",
    "TGL", "THALL", "TPLP", "TRG", "UBL", "UNITY", "WAVES", "YOUW",
}


PSX_SYMBOL_NAMES = {
    "ENGRO": "Engro Corporation Limited",
    "ENGROH": "Engro Holdings Limited",
    "FFC": "Fauji Fertilizer Company Limited",
    "FFBL": "Fauji Fertilizer Bin Qasim Limited",
    "HUBC": "Hub Power Company Limited",
    "LUCK": "Lucky Cement Limited",
    "MARI": "Mari Energies Limited",
    "MEBL": "Meezan Bank Limited",
    "SYS": "Systems Limited",
    "SSGC": "Sui Southern Gas Company Limited",
    "SNGP": "Sui Northern Gas Pipelines Limited",
    "OGDC": "Oil & Gas Development Company Limited",
    "PPL": "Pakistan Petroleum Limited",
    "PSO": "Pakistan State Oil Company Limited",
    "UBL": "United Bank Limited",
    "HBL": "Habib Bank Limited",
    "MCB": "MCB Bank Limited",
    "BAFL": "Bank Alfalah Limited",
    "BAHL": "Bank AL Habib Limited",
    "MTL": "Millat Tractors Limited",
    "MLCF": "Maple Leaf Cement Factory Limited",
    "DGKC": "D.G. Khan Cement Company Limited",
    "FCCL": "Fauji Cement Company Limited",
    "KOHC": "Kohat Cement Company Limited",
    "TRG": "TRG Pakistan Limited",
    "AIRLINK": "Air Link Communication Limited",
}


BROKER_NAME_KEYWORDS = [
    "js global",
    "cdc",
    "nccpl",
    "akd",
    "topline",
    "bma",
    "foundation securities",
    "arif habib",
    "alfalah securities",
    "broker",
    "account",
    "client",
]

BROKER_PORTFOLIO_KEYWORDS = [
    "portfolio",
    "holding",
    "holdings",
    "custody",
    "shares",
    "quantity",
    "symbol",
    "scrip",
    "security",
    "market value",
    "current price",
    "average price",
    "avg price",
    "cost price",
]




def read_portfolio_file(file_path):
    """
    Read supported file and return preview rows.

    Returns dictionary:
    {
        "success": bool,
        "file_type": str,
        "rows": list[dict],
        "errors": list[str],
        "warnings": list[str],
        "raw_text": str
    }
    """

    path = Path(file_path)

    result = {
        "success": False,
        "file_type": "",
        "rows": [],
        "errors": [],
        "warnings": [],
        "raw_text": "",
    }

    if not path.exists():
        result["errors"].append("Selected file does not exist.")
        return result

    extension = path.suffix.lower()
    result["file_type"] = extension

    if extension not in SUPPORTED_EXTENSIONS:
        result["errors"].append(
            "Unsupported file type. Supported: "
            + ", ".join(SUPPORTED_EXTENSIONS)
        )
        return result

    try:

        result["warnings"].append(f"Smart Import Service: {SMART_IMPORT_SERVICE_VERSION}")

        if extension == ".csv":
            rows = read_csv_file(path)
            result["rows"] = extract_holdings_from_rows(rows)

        elif extension in [".xlsx", ".xlsm"]:
            rows = read_excel_file(path)
            result["rows"] = extract_holdings_from_rows(rows)

        elif extension == ".pdf":
            text = extract_text_from_pdf(path)
            result["raw_text"] = text

            mixed_warning = detect_mixed_statement_warning(text)
            if mixed_warning:
                result["warnings"].append(mixed_warning)

            result["rows"] = extract_holdings_from_text(text)

        elif extension in [".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"]:
            text = extract_text_from_image(path)
            result["raw_text"] = text
            result["rows"] = extract_holdings_from_text(text)

        result["rows"] = clean_preview_rows(result["rows"])

        if result["rows"]:
            result["success"] = True
        else:
            result["warnings"].append(
                "No valid portfolio rows were detected. "
                "For PDF, make sure it is text-based and not scanned. "
                "CSV/Excel is most accurate."
            )

    except Exception as e:
        result["errors"].append(str(e))

    return result


def detect_mixed_statement_warning(text):
    """
    Return warning when PDF contains non-PSX statements.

    Current Smart Portfolio Import imports PSX holdings only.
    Mutual Funds/MTPF statements should be handled by a separate importer.
    """

    lowered = str(text or "").lower()

    has_al_meezan = (
        "al meezan investment management" in lowered
        or "meezan islamic fund" in lowered
        or "kse meezan index fund" in lowered
        or "mtpf" in lowered
    )

    has_psx_trades = "t+1 buy" in lowered or "t+1 sell" in lowered

    if has_al_meezan and has_psx_trades:
        return (
            "Mixed statement detected: PSX transactions plus Al Meezan "
            "Mutual Funds/MTPF pages. Smart Portfolio Import will import PSX "
            "holdings only. Mutual Fund/MTPF pages should be imported through "
            "a separate Mutual Fund/MTPF importer."
        )

    if has_al_meezan and not has_psx_trades:
        return (
            "Al Meezan Mutual Funds/MTPF statement detected. "
            "This file is not a PSX stock portfolio statement. "
            "Use Mutual Fund/MTPF importer when available."
        )

    return ""



def read_csv_file(path):
    """
    Read CSV file into list of dictionaries.
    """

    encodings = ["utf-8-sig", "utf-8", "cp1252", "latin1"]

    last_error = None

    for encoding in encodings:

        try:
            with open(path, "r", newline="", encoding=encoding) as file:
                sample = file.read(4096)
                file.seek(0)

                dialect = detect_csv_dialect(sample)

                reader = csv.DictReader(file, dialect=dialect)

                rows = []

                for row in reader:
                    rows.append(dict(row))

                if rows:
                    return rows

        except Exception as e:
            last_error = e

    raise ValueError(f"Could not read CSV file. {last_error}")


def detect_csv_dialect(sample):
    """
    Detect CSV delimiter safely.
    """

    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")

    except Exception:
        return csv.excel


def read_excel_file(path):
    """
    Read first useful Excel sheet into list of dictionaries.
    """

    try:
        from openpyxl import load_workbook

    except Exception as e:
        raise ImportError(
            "openpyxl is required to read Excel files. "
            "Install using: pip install openpyxl"
        ) from e

    workbook = load_workbook(str(path), data_only=True, read_only=True)

    best_rows = []

    for sheet in workbook.worksheets:

        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            continue

        extracted = rows_to_dicts(rows)

        if len(extracted) > len(best_rows):
            best_rows = extracted

    workbook.close()

    return best_rows


def rows_to_dicts(rows):
    """
    Convert Excel rows into dictionaries by detecting header row.
    """

    header_index = detect_header_row(rows)

    if header_index is None:
        return []

    headers = rows[header_index]
    normalized_headers = []

    for value in headers:
        text = normalize_header(value)

        if not text:
            text = f"column_{len(normalized_headers) + 1}"

        normalized_headers.append(text)

    dict_rows = []

    for row in rows[header_index + 1:]:

        if row_is_empty(row):
            continue

        item = {}

        for index, value in enumerate(row):

            if index >= len(normalized_headers):
                continue

            item[normalized_headers[index]] = value

        dict_rows.append(item)

    return dict_rows


def detect_header_row(rows):
    """
    Detect the row that most likely contains portfolio headers.
    """

    best_index = None
    best_score = 0

    for index, row in enumerate(rows[:30]):

        score = 0

        normalized_values = [
            normalize_header(value)
            for value in row
        ]

        for value in normalized_values:

            if value in SYMBOL_HEADERS:
                score += 5

            if value in NAME_HEADERS:
                score += 1

            if value in SHARES_HEADERS:
                score += 5

            if value in AVG_PRICE_HEADERS:
                score += 4

            if value in CURRENT_PRICE_HEADERS:
                score += 4

            if value in INVESTMENT_HEADERS:
                score += 2

            if value in CURRENT_VALUE_HEADERS:
                score += 2

        if score > best_score:
            best_score = score
            best_index = index

    if best_score >= 8:
        return best_index

    return None


def extract_text_from_pdf(path):
    """
    Extract text from text-based PDF.

    Tries:
    1. pypdf
    2. PyPDF2
    3. pdfplumber
    4. pymupdf / fitz

    Install recommended:
        pip install pypdf
    """

    errors = []
    text_parts = []

    # 1) pypdf
    try:
        from pypdf import PdfReader

        reader = PdfReader(str(path))

        for page in reader.pages:
            text_parts.append(page.extract_text() or "")

        text = "\n".join(text_parts).strip()

        if text:
            return normalize_pdf_text(text)

    except Exception as e:
        errors.append(f"pypdf: {e}")

    # 2) PyPDF2
    try:
        from PyPDF2 import PdfReader

        reader = PdfReader(str(path))

        text_parts = []

        for page in reader.pages:
            text_parts.append(page.extract_text() or "")

        text = "\n".join(text_parts).strip()

        if text:
            return normalize_pdf_text(text)

    except Exception as e:
        errors.append(f"PyPDF2: {e}")

    # 3) pdfplumber
    try:
        import pdfplumber

        text_parts = []

        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                text_parts.append(page.extract_text() or "")

        text = "\n".join(text_parts).strip()

        if text:
            return normalize_pdf_text(text)

    except Exception as e:
        errors.append(f"pdfplumber: {e}")

    # 4) PyMuPDF / fitz
    try:
        import fitz

        text_parts = []

        document = fitz.open(str(path))

        for page in document:
            text_parts.append(page.get_text("text") or "")

        document.close()

        text = "\n".join(text_parts).strip()

        if text:
            return normalize_pdf_text(text)

    except Exception as e:
        errors.append(f"pymupdf: {e}")

    raise ValueError(
        "Could not extract text from PDF.\n\n"
        "Please install PDF text support:\n"
        "pip install pypdf\n\n"
        "If this is a scanned PDF, OCR will be required.\n\n"
        "Details:\n"
        + "\n".join(errors[-4:])
    )


def normalize_pdf_text(text):
    """
    Normalize PDF extracted text.
    """

    text = str(text or "")
    text = text.replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n\s+\n", "\n", text)

    return text.strip()


def extract_text_from_image(path):
    """
    Extract text from JPG/PNG screenshot using OCR.

    Requirements:
    1. pip install pillow pytesseract
    2. Install Tesseract OCR on Windows.
       Common path:
       C:\\Program Files\\Tesseract-OCR\\tesseract.exe
    """

    try:
        from PIL import Image, ImageOps, ImageFilter
        import pytesseract

    except Exception as e:
        raise ImportError(
            "Image import requires OCR Python packages.\n\n"
            "Please run this command:\n"
            "pip install pillow pytesseract\n\n"
            "Then install Tesseract OCR for Windows."
        ) from e

    configure_tesseract_if_available(pytesseract)
    validate_tesseract_language_data()

    try:
        image = Image.open(str(path))

    except Exception as e:
        raise ValueError(f"Could not open selected image file. {e}") from e

    ocr_texts = []

    # Try original image first.
    try:
        ocr_texts.append(
            pytesseract.image_to_string(
                image,
                lang="eng",
                config=get_tesseract_config(6)
            )
        )
    except Exception as e:
        raise RuntimeError(
            "Tesseract OCR is not working.\n\n"
            "Please install Tesseract OCR on Windows and restart the app.\n"
            "Expected path usually:\n"
            "C:\\Program Files\\Tesseract-OCR\\tesseract.exe\n\n"
            f"TESSDATA_PREFIX: {os.environ.get('TESSDATA_PREFIX', '')}\n"
            f"Details: {e}"
        ) from e

    # Try preprocessed image for screenshots.
    try:
        processed = preprocess_image_for_ocr(image)
        ocr_texts.append(
            pytesseract.image_to_string(
                processed,
                lang="eng",
                config=get_tesseract_config(6)
            )
        )
    except Exception:
        pass

    # Try table-like OCR mode.
    try:
        processed = preprocess_image_for_ocr(image)
        ocr_texts.append(
            pytesseract.image_to_string(
                processed,
                lang="eng",
                config=get_tesseract_config(11)
            )
        )
    except Exception:
        pass

    text = "\n".join([t for t in ocr_texts if t and t.strip()]).strip()

    if not text:
        raise ValueError(
            "No readable text detected from image.\n\n"
            "Please use a clearer/high-resolution screenshot, "
            "or use CSV/Excel for best accuracy."
        )

    return normalize_pdf_text(text)


def configure_tesseract_if_available(pytesseract):
    """
    Configure pytesseract with common Windows install paths.

    This version sets:
    - tesseract_cmd
    - TESSDATA_PREFIX without quotes
    - PATH fallback
    """

    possible_paths = [
        Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe"),
        Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"),
    ]

    for tesseract_path in possible_paths:

        if not tesseract_path.exists():
            continue

        pytesseract.pytesseract.tesseract_cmd = str(tesseract_path)

        install_folder = tesseract_path.parent
        tessdata_path = install_folder / "tessdata"

        if tessdata_path.exists():
            # No quotes here. Tesseract reads this as an environment variable.
            os.environ["TESSDATA_PREFIX"] = str(tessdata_path)

        # Add install folder to PATH for subprocess dependencies.
        current_path = os.environ.get("PATH", "")

        if str(install_folder) not in current_path:
            os.environ["PATH"] = str(install_folder) + os.pathsep + current_path

        return


def get_tesseract_config(psm=6):
    """
    Return Tesseract OCR config.

    Do not include --tessdata-dir here; TESSDATA_PREFIX handles it.
    """

    return f"--psm {psm}"

def validate_tesseract_language_data():
    """
    Validate English OCR language data exists.
    """

    possible_eng_paths = [
        Path(r"C:\Program Files\Tesseract-OCR\tessdata\eng.traineddata"),
        Path(r"C:\Program Files (x86)\Tesseract-OCR\tessdata\eng.traineddata"),
    ]

    for eng_path in possible_eng_paths:

        if eng_path.exists():
            return True

    raise RuntimeError(
        "Tesseract English language data is missing.\n\n"
        "Required file:\n"
        "C:\\Program Files\\Tesseract-OCR\\tessdata\\eng.traineddata\n\n"
        "Fix:\n"
        "1. Re-run Tesseract installer\n"
        "2. Select English language data during installation\n"
        "3. Or copy eng.traineddata into the tessdata folder"
    )


def preprocess_image_for_ocr(image):
    """
    Improve screenshot readability for OCR.
    """

    from PIL import ImageOps, ImageFilter

    image = image.convert("L")

    # Enlarge small screenshots.
    width, height = image.size

    if width < 1600:
        scale = 1600 / max(width, 1)
        new_size = (
            int(width * scale),
            int(height * scale),
        )
        image = image.resize(new_size)

    image = ImageOps.autocontrast(image)
    image = image.filter(ImageFilter.SHARPEN)

    # Simple threshold.
    image = image.point(lambda pixel: 255 if pixel > 180 else 0)

    return image


def extract_holdings_from_rows(rows):
    """
    Extract holdings from CSV/Excel dictionary rows.
    """

    if not rows:
        return []

    cleaned_rows = []
    normalized_rows = []

    for row in rows:

        normalized_row = {}

        for key, value in row.items():
            normalized_key = normalize_header(key)
            normalized_row[normalized_key] = value

        normalized_rows.append(normalized_row)

    columns = detect_columns(normalized_rows)

    for row in normalized_rows:

        symbol = get_row_value(row, columns.get("symbol"))
        share_name = get_row_value(row, columns.get("share_name"))
        shares = get_row_value(row, columns.get("shares"))
        avg_price = get_row_value(row, columns.get("avg_price"))
        current_price = get_row_value(row, columns.get("current_price"))

        investment_value = get_row_value(row, columns.get("investment"))
        current_value = get_row_value(row, columns.get("current_value"))

        symbol = normalize_symbol(symbol)
        share_name = clean_share_name(share_name)
        shares = parse_number(shares)
        avg_price = parse_number(avg_price)
        current_price = parse_number(current_price)

        investment_value = parse_number(investment_value)
        current_value = parse_number(current_value)

        if avg_price <= 0 and investment_value > 0 and shares > 0:
            avg_price = investment_value / shares

        if current_price <= 0 and current_value > 0 and shares > 0:
            current_price = current_value / shares

        if not is_valid_symbol(symbol):
            continue

        if shares <= 0:
            continue

        cleaned_rows.append({
            "symbol": symbol,
            "share_name": share_name,
            "shares": round(shares, 4),
            "avg_price": round(avg_price, 4),
            "current_price": round(current_price, 4),
            "investment_value": round(shares * avg_price, 2),
            "current_value": round(shares * current_price, 2),
            "source": "Structured File",
            "confidence": calculate_confidence(symbol, shares, avg_price, current_price),
            "remarks": "",
        })

    return cleaned_rows


def detect_columns(rows):
    """
    Detect portfolio columns from normalized rows.
    """

    all_headers = set()

    for row in rows:
        for key in row.keys():
            all_headers.add(key)

    return {
        "symbol": find_best_header(all_headers, SYMBOL_HEADERS),
        "share_name": find_best_header(all_headers, NAME_HEADERS),
        "shares": find_best_header(all_headers, SHARES_HEADERS),
        "avg_price": find_best_header(all_headers, AVG_PRICE_HEADERS),
        "current_price": find_best_header(all_headers, CURRENT_PRICE_HEADERS),
        "investment": find_best_header(all_headers, INVESTMENT_HEADERS),
        "current_value": find_best_header(all_headers, CURRENT_VALUE_HEADERS),
    }


def find_best_header(headers, candidates):
    """
    Find exact or partial header match.
    """

    for candidate in candidates:
        if candidate in headers:
            return candidate

    for header in headers:
        for candidate in candidates:
            if candidate in header or header in candidate:
                return header

    return None


def extract_holdings_from_text(text):
    """
    Extract holdings from PDF/image raw text.

    Priority:
    1. Broker account transaction statement:
       T+1 Buy/Sell SYMBOL Qty @ Price
    2. Broker/portfolio table exact extractor.
    3. Focused table extractor.
    4. Common PSX symbol fallback for screenshots.
    """

    if not text:
        return []

    lines = get_clean_text_lines(text)

    # JS InvestPro mobile screenshots show:
    # SYMBOL current_price
    # qty shares @ avg_price
    js_investpro_rows = extract_js_investpro_screenshot_rows(lines)

    if js_investpro_rows:
        return js_investpro_rows

    # Transaction ledger statements do not have a current holdings table.
    # Build net holdings from Buy/Sell transactions across all extracted pages.
    transaction_rows = extract_broker_transaction_statement_rows(lines)

    if transaction_rows:
        return transaction_rows

    exact_rows = extract_broker_exact_rows(lines)

    if exact_rows:
        return exact_rows

    rows = []

    table_lines = extract_portfolio_table_lines(lines)

    if table_lines:
        candidate_lines = table_lines
    else:
        candidate_lines = build_strict_candidate_lines(lines)

    for line in candidate_lines:

        parsed = parse_holding_line(line)

        if parsed:
            rows.append(parsed)

    if rows:
        return dedupe_exact_rows(rows)

    # Last fallback for JPG/PNG screenshots where OCR table headers are weak.
    loose_rows = extract_common_symbol_screenshot_rows(lines)

    return dedupe_exact_rows(loose_rows)


def extract_common_symbol_screenshot_rows(lines):
    """
    Fallback for screenshots where OCR does not preserve headers.

    It only accepts rows containing a known PSX symbol and at least 3 numeric
    values after that symbol. This prevents random words like HOUSE from being
    imported.
    """

    rows = []

    for line in lines:

        parsed = parse_common_symbol_screenshot_line(line)

        if parsed:
            rows.append(parsed)

    return rows


def parse_common_symbol_screenshot_line(line):
    """
    Parse screenshot line by known PSX symbol.

    Examples:
    FFC Fauji Fertilizer Company Limited 37 556.95 556.95
    MEBL 34 490.88 490.88
    """

    line = clean_ocr_line(line)

    if not line:
        return None

    if is_noise_line(line):
        return None

    tokens = line.split()

    if len(tokens) < 4:
        return None

    symbol = ""

    for token in tokens[:6]:

        possible = normalize_symbol(token)

        if possible in COMMON_PSX_SYMBOLS:
            symbol = possible
            break

    if not symbol:
        return None

    numbers = extract_numbers_from_text_after_symbol(line, symbol)

    if len(numbers) < 3:
        return None

    # For screenshots, the first 3 useful numbers after symbol are usually:
    # shares, avg/current price, current/market price.
    shares = numbers[0]
    avg_price = numbers[1]
    current_price = numbers[2]

    if shares <= 0 or avg_price <= 0 or current_price <= 0:
        return None

    if shares > 100000000:
        return None

    share_name = extract_share_name_from_line(line, symbol)

    confidence = calculate_confidence(symbol, shares, avg_price, current_price)

    if share_name:
        confidence += 5

    if confidence > 100:
        confidence = 100

    return {
        "symbol": symbol,
        "share_name": share_name or PSX_SYMBOL_NAMES.get(symbol, ""),
        "shares": round(shares, 4),
        "avg_price": round(avg_price, 4),
        "current_price": round(current_price, 4),
        "investment_value": round(shares * avg_price, 2),
        "current_value": round(shares * current_price, 2),
        "source": "Image/PDF Screenshot Fallback",
        "confidence": confidence,
        "remarks": "Fallback extraction. Please verify row before importing.",
    }


def extract_js_investpro_screenshot_rows(lines):
    """
    Extract holdings from JS InvestPro mobile app screenshots.

    Common OCR format:
    LUCK 470.80
    16 shares @ 437.61
    Investment Value Current Value
    PKR 7,001.76 PKR 7,532.80

    MARI 673.60
    48 shares @ 661.36
    """

    if not lines:
        return []

    joined = " ".join(lines).lower()

    # Only activate this parser for mobile portfolio/watchlist screenshots.
    if (
        "investpro" not in joined
        and "watchlist" not in joined
        and "shares @" not in joined
        and "investment value current value" not in joined
    ):
        return []

    rows = []

    for index, line in enumerate(lines):

        symbol_info = parse_js_symbol_current_price_line(line)

        if not symbol_info:
            continue

        symbol = symbol_info["symbol"]
        current_price = symbol_info["current_price"]

        share_info = find_js_shares_avg_price(lines, index + 1, max_lookahead=5)

        if not share_info:
            continue

        shares = share_info["shares"]
        avg_price = share_info["avg_price"]

        # Try to read investment/current value from nearby lines.
        values = find_js_investment_current_values(lines, index, max_lookahead=14)

        investment_value = shares * avg_price
        current_value = shares * current_price

        if values:
            investment_value = values.get("investment_value", investment_value)
            current_value = values.get("current_value", current_value)

            if shares > 0 and current_value > 0:
                derived_current_price = current_value / shares

                if derived_current_price > 0:
                    current_price = derived_current_price

        rows.append({
            "symbol": symbol,
            "share_name": PSX_SYMBOL_NAMES.get(symbol, ""),
            "shares": round(shares, 4),
            "avg_price": round(avg_price, 4),
            "current_price": round(current_price, 4),
            "investment_value": round(shares * avg_price, 2),
            "current_value": round(shares * current_price, 2),
            "source": "JS InvestPro Screenshot",
            "confidence": 96,
            "remarks": "Extracted from JS InvestPro screenshot. Please verify before importing.",
        })

    return dedupe_exact_rows(rows)


def parse_js_symbol_current_price_line(line):
    """
    Parse line containing symbol and current price.

    Examples:
    ry LUCK 470.80
    6 MARI 673.60
    em SYS 148.66
    """

    line = clean_ocr_line(line)

    if not line:
        return None

    if is_noise_line(line):
        return None

    tokens = line.split()

    if len(tokens) < 2:
        return None

    for pos, token in enumerate(tokens[:6]):

        symbol = normalize_symbol(token)

        if symbol not in COMMON_PSX_SYMBOLS:
            continue

        numbers_after = extract_numbers_from_tokens(tokens[pos + 1:])

        if not numbers_after:
            continue

        current_price = numbers_after[0]

        if current_price <= 0:
            continue

        # Reject impossible current price from index/volume lines.
        if current_price > 100000:
            continue

        return {
            "symbol": symbol,
            "current_price": current_price,
        }

    return None


def find_js_shares_avg_price(lines, start_index, max_lookahead=5):
    """
    Find 'N shares @ price' line after symbol line.
    """

    end_index = min(len(lines), start_index + max_lookahead)

    pattern = re.compile(
        r"([0-9,]+(?:\.[0-9]+)?)\s+shares?\s*@\s*([0-9,]+(?:\.[0-9]+)?)",
        re.IGNORECASE
    )

    for index in range(start_index, end_index):

        line = clean_ocr_line(lines[index])
        match = pattern.search(line)

        if not match:
            continue

        shares = parse_number(match.group(1))
        avg_price = parse_number(match.group(2))

        if shares > 0 and avg_price > 0:
            return {
                "shares": shares,
                "avg_price": avg_price,
            }

    return None


def find_js_investment_current_values(lines, start_index, max_lookahead=14):
    """
    Find Investment Value / Current Value numbers near the holding block.
    """

    end_index = min(len(lines), start_index + max_lookahead)

    for index in range(start_index, end_index):

        line = clean_ocr_line(lines[index]).lower()

        if "investment value" not in line or "current value" not in line:
            continue

        # Usually values are on the next line.
        for value_index in range(index + 1, min(len(lines), index + 4)):

            value_line = clean_ocr_line(lines[value_index])
            numbers = extract_all_numbers_from_text(value_line)

            if len(numbers) >= 2:
                return {
                    "investment_value": numbers[0],
                    "current_value": numbers[1],
                }

    return None


def extract_all_numbers_from_text(text):
    """
    Extract all numbers from a text line.
    """

    text = str(text or "")
    text = text.replace(",", "")
    text = text.replace("PKR", "")
    text = text.replace("pkr", "")
    text = text.replace("pxr", "")
    text = text.replace("pwr", "")
    text = text.replace("par", "")
    text = text.replace("per", "")

    matches = re.findall(r"-?\d+(?:\.\d+)?", text)

    numbers = []

    for match in matches:
        try:
            numbers.append(float(match))
        except Exception:
            pass

    return numbers


def extract_broker_transaction_statement_rows(lines):
    """
    Extract net PSX holdings from broker account statement transactions.

    Handles lines like:
    May 20,2026 T+1 Buy MEBL 1 @ 463.96000000 Comm Amt ...
    Jun 02,2026 T+1 Sell SSGC 1 @ 27.30000000 Comm Amt ...

    Output:
    - Net shares
    - Weighted average buy price
    - Current price estimated from latest transaction price
    """

    trades = []

    for line in lines:
        trade = parse_broker_trade_line(line)

        if trade:
            trades.append(trade)

    if not trades:
        return []

    positions = {}

    for trade in trades:

        symbol = trade["symbol"]
        side = trade["side"]
        quantity = trade["quantity"]
        price = trade["price"]

        if symbol not in positions:
            positions[symbol] = {
                "shares": 0.0,
                "cost": 0.0,
                "last_price": 0.0,
                "buy_count": 0,
                "sell_count": 0,
            }

        position = positions[symbol]

        if side == "BUY":
            position["shares"] += quantity
            position["cost"] += quantity * price
            position["last_price"] = price
            position["buy_count"] += 1

        elif side == "SELL":

            if position["shares"] > 0:
                average_cost = position["cost"] / position["shares"]

                sell_quantity = quantity

                if sell_quantity > position["shares"]:
                    sell_quantity = position["shares"]

                position["shares"] -= sell_quantity
                position["cost"] -= average_cost * sell_quantity

                if position["shares"] < 0:
                    position["shares"] = 0

                if position["cost"] < 0:
                    position["cost"] = 0

            position["last_price"] = price
            position["sell_count"] += 1

    rows = []

    for symbol, position in positions.items():

        shares = position["shares"]

        if shares <= 0:
            continue

        if shares < 0.0001:
            continue

        if position["cost"] > 0:
            avg_price = position["cost"] / shares
        else:
            avg_price = position["last_price"]

        current_price = position["last_price"] or avg_price

        rows.append({
            "symbol": symbol,
            "share_name": PSX_SYMBOL_NAMES.get(symbol, ""),
            "shares": round(shares, 4),
            "avg_price": round(avg_price, 4),
            "current_price": round(current_price, 4),
            "investment_value": round(shares * avg_price, 2),
            "current_value": round(shares * current_price, 2),
            "source": "Broker Transaction Statement",
            "confidence": 95,
            "remarks": (
                "Calculated from Buy/Sell transaction statement. "
                "Current price is estimated from latest transaction price; "
                "please update current market price if needed."
            ),
        })

    rows.sort(key=lambda item: item["symbol"])

    return rows


def parse_broker_trade_line(line):
    """
    Parse one buy/sell transaction line.
    """

    line = clean_ocr_line(line)

    if not line:
        return None

    pattern = re.compile(
        r"T\s*\+\s*1\s+"
        r"(Buy|Sell)\s+"
        r"([A-Z0-9]+)\s+"
        r"([0-9,]+(?:\.[0-9]+)?)\s+"
        r"@\s*"
        r"([0-9,]+(?:\.[0-9]+)?)",
        re.IGNORECASE
    )

    match = pattern.search(line)

    if not match:
        return None

    side = match.group(1).strip().upper()
    symbol = normalize_symbol(match.group(2))
    quantity = parse_number(match.group(3))
    price = parse_number(match.group(4))

    if not is_valid_symbol(symbol):
        return None

    if quantity <= 0 or price <= 0:
        return None

    if symbol not in COMMON_PSX_SYMBOLS:
        # Trade statements are strict; avoid false OCR symbols.
        return None

    return {
        "side": side,
        "symbol": symbol,
        "quantity": quantity,
        "price": price,
    }


def extract_broker_exact_rows(lines):
    """
    Extract portfolio rows with broker-style table detection.

    This is stricter than generic OCR parsing:
    - finds a holdings/portfolio table area
    - accepts rows that start with a likely PSX symbol
    - uses first 3 numeric columns as Shares, Avg Price, Current Price
    """

    if not lines:
        return []

    table_blocks = find_broker_portfolio_blocks(lines)
    rows = []

    for block in table_blocks:

        for line in block:
            parsed = parse_broker_exact_row(line, inside_table=True)

            if parsed:
                rows.append(parsed)

    if rows:
        return dedupe_exact_rows(rows)

    # Very strict fallback for clean screenshots/PDFs without a detectable header.
    for line in lines:
        parsed = parse_broker_exact_row(line, inside_table=False)

        if parsed:
            rows.append(parsed)

    return dedupe_exact_rows(rows)


def find_broker_portfolio_blocks(lines):
    """
    Find blocks likely containing portfolio holdings table.
    """

    blocks = []
    current_block = []
    in_block = False
    non_row_count = 0

    for index, line in enumerate(lines):

        lowered = line.lower()

        if is_broker_portfolio_header(line):
            in_block = True
            current_block = []
            non_row_count = 0
            continue

        if not in_block:
            continue

        if is_portfolio_block_stop(line):

            if current_block:
                blocks.append(current_block)

            in_block = False
            current_block = []
            non_row_count = 0
            continue

        if looks_like_broker_exact_row(line):
            current_block.append(line)
            non_row_count = 0
        else:
            non_row_count += 1

        # Stop block if several non-row lines appear after holdings started.
        if current_block and non_row_count >= 4:
            blocks.append(current_block)
            in_block = False
            current_block = []
            non_row_count = 0

    if current_block:
        blocks.append(current_block)

    return blocks


def is_broker_portfolio_header(line):
    """
    Detect table header for holdings/portfolio.
    """

    lowered = clean_ocr_line(line).lower()

    if not lowered:
        return False

    score = 0

    for keyword in BROKER_PORTFOLIO_KEYWORDS:
        if keyword in lowered:
            score += 1

    has_symbol = any(word in lowered for word in ["symbol", "scrip", "security", "stock"])
    has_qty = any(word in lowered for word in ["qty", "quantity", "shares", "holding", "holdings"])
    has_price = any(word in lowered for word in ["price", "rate", "ltp", "market", "avg", "average", "cost"])

    if has_symbol and has_qty and has_price:
        score += 5

    return score >= 5


def is_portfolio_block_stop(line):
    """
    Detect where a portfolio table likely ends.
    """

    lowered = clean_ocr_line(line).lower()

    stop_phrases = [
        "grand total",
        "portfolio total",
        "total value",
        "cash balance",
        "ledger",
        "transaction",
        "dividend",
        "capital gain",
        "disclaimer",
        "note:",
        "page ",
        "generated",
        "signature",
    ]

    return any(phrase in lowered for phrase in stop_phrases)


def looks_like_broker_exact_row(line):
    """
    Strict check for broker holding row.
    """

    line = clean_ocr_line(line)

    if not line or is_noise_line(line):
        return False

    tokens = line.split()

    if len(tokens) < 4:
        return False

    symbol = normalize_symbol(tokens[0])

    if not is_valid_symbol(symbol):
        return False

    numbers = extract_numbers_from_text_after_symbol(line, symbol)

    if len(numbers) < 3:
        return False

    if line_contains_mostly_date_time(line):
        return False

    return True


def parse_broker_exact_row(line, inside_table=False):
    """
    Parse broker table row.

    Expected common formats:
    SYMBOL Company Name Shares AvgPrice CurrentPrice
    SYMBOL Shares AvgPrice CurrentPrice
    SYMBOL Company Name Qty CostPrice MarketPrice MarketValue ...
    """

    line = clean_ocr_line(line)

    if not looks_like_broker_exact_row(line):
        return None

    tokens = line.split()
    symbol = normalize_symbol(tokens[0])

    if not inside_table and symbol not in COMMON_PSX_SYMBOLS:
        # Outside a detected table, avoid random OCR words.
        return None

    numbers = extract_numbers_from_text_after_symbol(line, symbol)

    if len(numbers) < 3:
        return None

    # For broker tables, first three numeric values after name are normally:
    # shares, average/cost price, current/market price.
    shares = numbers[0]
    avg_price = numbers[1]
    current_price = numbers[2]

    if shares <= 0:
        return None

    if avg_price < 0 or current_price < 0:
        return None

    if shares > 100000000:
        return None

    share_name = extract_share_name_from_line(line, symbol)

    confidence = calculate_confidence(symbol, shares, avg_price, current_price)

    if inside_table:
        confidence += 5

    if share_name:
        confidence += 5

    if confidence > 100:
        confidence = 100

    return {
        "symbol": symbol,
        "share_name": share_name,
        "shares": round(shares, 4),
        "avg_price": round(avg_price, 4),
        "current_price": round(current_price, 4),
        "investment_value": round(shares * avg_price, 2),
        "current_value": round(shares * current_price, 2),
        "source": "Broker Exact OCR/PDF",
        "confidence": confidence,
        "remarks": "Broker-style exact extraction. Please verify before importing.",
    }


def extract_numbers_from_text_after_symbol(line, symbol):
    """
    Extract numeric values after symbol.
    """

    line = clean_ocr_line(line)
    symbol = normalize_symbol(symbol)

    tokens = line.split()

    after_symbol = []
    found = False

    for token in tokens:

        if not found:

            if normalize_symbol(token) == symbol:
                found = True

            continue

        after_symbol.append(token)

    return extract_numbers_from_tokens(after_symbol)


def dedupe_exact_rows(rows):
    """
    Remove duplicate extracted rows.
    """

    result = []
    seen = set()

    for row in rows:

        key = (
            row.get("symbol"),
            round(parse_number(row.get("shares", 0)), 4),
            round(parse_number(row.get("avg_price", 0)), 4),
            round(parse_number(row.get("current_price", 0)), 4),
        )

        if key in seen:
            continue

        seen.add(key)
        result.append(row)

    return result


def get_clean_text_lines(text):
    """
    Return cleaned lines from OCR/PDF text.
    """

    raw_lines = str(text or "").replace("\r", "\n").splitlines()

    lines = []

    for line in raw_lines:

        line = clean_ocr_line(line)

        if not line:
            continue

        if is_noise_line(line):
            continue

        lines.append(line)

    return lines


def clean_ocr_line(line):
    """
    Clean one OCR/PDF text line.
    """

    line = str(line or "").strip()

    if not line:
        return ""

    replacements = {
        "|": " ",
        "—": "-",
        "–": "-",
        "•": " ",
        "_": " ",
        "\t": " ",
    }

    for old, new in replacements.items():
        line = line.replace(old, new)

    line = re.sub(r"\s+", " ", line)
    return line.strip()


def is_noise_line(line):
    """
    Identify obvious non-portfolio text lines.
    """

    lowered = str(line or "").lower()

    if not lowered:
        return True

    for keyword in NOISE_LINE_KEYWORDS:
        if keyword in lowered:
            return True

    # Ignore standalone dates/times lines.
    if re.fullmatch(r"[\d\s:/\-.apmAPM]+", line):
        return True

    return False


def extract_portfolio_table_lines(lines):
    """
    Extract lines likely belonging to a portfolio holdings table.

    Starts after a header line with symbol/qty/price keywords.
    Stops when summary/total/noise section begins.
    """

    if not lines:
        return []

    header_index = find_portfolio_header_index(lines)

    if header_index is None:
        return []

    table_lines = []

    for line in lines[header_index + 1:]:

        lowered = line.lower()

        if any(keyword in lowered for keyword in PORTFOLIO_STOP_KEYWORDS):
            # Keep "total" as stop only after some rows have been collected.
            if table_lines:
                break

        if looks_like_portfolio_row(line):
            table_lines.append(line)
            continue

        # Allow split OCR rows: combine current line with next few later in fallback.
        if table_lines and len(table_lines) >= 1:
            # Stop after table if several non-row lines appear.
            pass

    # If direct table lines failed, try chunks after header.
    if not table_lines:

        remaining = lines[header_index + 1:]

        for index in range(len(remaining)):
            chunk = " ".join(remaining[index:index + 4])

            if looks_like_portfolio_row(chunk):
                table_lines.append(chunk)

    return table_lines


def find_portfolio_header_index(lines):
    """
    Find likely portfolio holdings table header.
    """

    best_index = None
    best_score = 0

    for index, line in enumerate(lines[:80]):

        lowered = line.lower()
        score = 0

        for keyword in PORTFOLIO_HEADER_KEYWORDS:
            if keyword in lowered:
                score += 1

        has_symbol_word = any(
            key in lowered
            for key in ["symbol", "scrip", "ticker", "security", "stock"]
        )

        has_quantity_word = any(
            key in lowered
            for key in ["qty", "quantity", "shares", "holding"]
        )

        has_price_word = any(
            key in lowered
            for key in ["price", "rate", "ltp", "market", "avg", "average"]
        )

        if has_symbol_word and has_quantity_word and has_price_word:
            score += 8

        if score > best_score:
            best_score = score
            best_index = index

    if best_score >= 8:
        return best_index

    return None


def build_strict_candidate_lines(lines):
    """
    Build candidate holding lines without reading every random OCR word.

    Fallback rules are intentionally strict:
    - Symbol must appear at start/near-start.
    - Line/chunk must contain at least 3 numeric values.
    """

    candidate_lines = []

    for line in lines:

        if looks_like_portfolio_row(line):
            candidate_lines.append(line)

    # Handle split OCR rows by joining nearby lines, but only when first line
    # begins with a possible stock symbol.
    for index, line in enumerate(lines):

        first_token = get_first_token(line)

        if not is_valid_symbol(first_token):
            continue

        chunk = " ".join(lines[index:index + 5])

        if looks_like_portfolio_row(chunk):
            candidate_lines.append(chunk)

    return candidate_lines


def looks_like_portfolio_row(line):
    """
    Check whether a line looks like a portfolio holding row.

    Required:
    - starts with or near-start contains valid stock symbol
    - at least 3 useful numeric values
    - not obvious noise/header
    """

    line = clean_ocr_line(line)

    if not line:
        return False

    lowered = line.lower()

    if is_noise_line(line):
        return False

    if any(
        phrase in lowered
        for phrase in [
            "symbol share",
            "shares avg",
            "current price",
            "portfolio import",
            "sample psx",
            "columns include",
        ]
    ):
        return False

    tokens = line.split()

    if len(tokens) < 4:
        return False

    possible_symbol = find_symbol_in_early_tokens(tokens)

    if not possible_symbol:
        return False

    numbers = extract_numbers_from_tokens(tokens)

    if len(numbers) < 3:
        return False

    # Avoid reading lines where numbers are date/time fragments only.
    if line_contains_mostly_date_time(line):
        return False

    return True


def find_symbol_in_early_tokens(tokens):
    """
    Find a likely stock symbol in first few tokens.
    """

    for token in tokens[:3]:

        symbol = normalize_symbol(token)

        if is_valid_symbol(symbol):
            return symbol

    return ""


def extract_numbers_from_tokens(tokens):
    """
    Extract positive numeric values from tokens.
    """

    numbers = []

    for token in tokens:
        number = parse_number(token)

        if number > 0:
            numbers.append(number)

    return numbers


def line_contains_mostly_date_time(line):
    """
    Detect date/time heavy lines to avoid WhatsApp/date false positives.
    """

    lowered = line.lower()

    if "am" in lowered or "pm" in lowered:
        return True

    if re.search(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", lowered):
        return True

    if re.search(r"\d{1,2}[:.]\d{2}", lowered):
        return True

    return False


def parse_holding_line(line):
    """
    Parse one possible portfolio holding line.
    """

    original_line = clean_ocr_line(line)

    if not original_line:
        return None

    if not looks_like_portfolio_row(original_line):
        return None

    line = original_line.replace(",", "")
    tokens = line.split()

    symbol = find_symbol_in_early_tokens(tokens)

    if not symbol:
        return None

    if symbol in ["HOUSE", "MOHALLA", "COLONY", "JHANG", "PAKISTAN"]:
        return None

    if symbol not in COMMON_PSX_SYMBOLS:
        return None

    numbers = extract_numbers_from_tokens(tokens)

    if len(numbers) < 3:
        return None

    # For holding rows with company name:
    # SYMBOL Company Name Shares AvgPrice CurrentPrice
    # Use last 3 numbers to avoid account/page/date numbers at front.
    shares = numbers[-3]
    avg_price = numbers[-2]
    current_price = numbers[-1]

    if shares <= 0:
        return None

    if avg_price < 0 or current_price < 0:
        return None

    # Sanity filters: qty normally not absurdly huge for small personal portfolio,
    # but keep flexible.
    if shares > 100000000:
        return None

    share_name = extract_share_name_from_line(line, symbol)

    return {
        "symbol": symbol,
        "share_name": share_name,
        "shares": round(shares, 4),
        "avg_price": round(avg_price, 4),
        "current_price": round(current_price, 4),
        "investment_value": round(shares * avg_price, 2),
        "current_value": round(shares * current_price, 2),
        "source": "PDF/Image Text",
        "confidence": calculate_text_row_confidence(
            symbol,
            share_name,
            shares,
            avg_price,
            current_price,
            original_line,
        ),
        "remarks": "Please verify extracted PDF/image row before importing.",
    }


def calculate_text_row_confidence(symbol, share_name, shares, avg_price, current_price, line):
    """
    Confidence for PDF/image extracted rows.
    """

    score = calculate_confidence(symbol, shares, avg_price, current_price) - 10

    if share_name:
        score += 5

    if looks_like_portfolio_row(line):
        score += 5

    if score < 0:
        score = 0

    if score > 100:
        score = 100

    return score


def get_first_token(line):
    """
    Return first token from line.
    """

    tokens = str(line or "").split()

    if not tokens:
        return ""

    return tokens[0]


def extract_share_name_from_line(line, symbol):
    """
    Extract share/company name between symbol and numeric columns.
    """

    line = str(line or "").strip()
    symbol = normalize_symbol(symbol)

    if not line or not symbol:
        return ""

    parts = line.split()

    cleaned = []

    found_symbol = False

    for token in parts:

        if not found_symbol:

            if normalize_symbol(token) == symbol:
                found_symbol = True

            continue

        if parse_number(token) > 0:
            break

        cleaned.append(token)

    return clean_share_name(" ".join(cleaned))


def clean_preview_rows(rows):
    """
    Clean duplicate rows and normalize confidence.
    """

    final_rows = []
    seen = set()

    for row in rows or []:

        symbol = normalize_symbol(row.get("symbol", ""))
        share_name = clean_share_name(row.get("share_name", ""))
        shares = parse_number(row.get("shares", 0))
        avg_price = parse_number(row.get("avg_price", 0))
        current_price = parse_number(row.get("current_price", 0))

        if not is_valid_symbol(symbol):
            continue

        if shares <= 0:
            continue

        key = (
            symbol,
            round(shares, 4),
            round(avg_price, 4),
            round(current_price, 4),
        )

        if key in seen:
            continue

        seen.add(key)

        confidence = int(row.get("confidence", 70))

        if confidence < 0:
            confidence = 0

        if confidence > 100:
            confidence = 100

        final_rows.append({
            "symbol": symbol,
            "share_name": share_name,
            "shares": round(shares, 4),
            "avg_price": round(avg_price, 4),
            "current_price": round(current_price, 4),
            "investment_value": round(shares * avg_price, 2),
            "current_value": round(shares * current_price, 2),
            "source": row.get("source", ""),
            "confidence": confidence,
            "remarks": row.get("remarks", ""),
        })

    return final_rows


def calculate_confidence(symbol, shares, avg_price, current_price):
    """
    Return simple confidence score.
    """

    score = 0

    if is_valid_symbol(symbol):
        score += 35

    if shares > 0:
        score += 25

    if avg_price > 0:
        score += 20

    if current_price > 0:
        score += 20

    return score


def get_row_value(row, key):
    """
    Get row value safely.
    """

    if not key:
        return ""

    return row.get(key, "")


def normalize_header(value):
    """
    Normalize column header text.
    """

    text = str(value or "").strip().lower()
    text = text.replace("\n", " ")
    text = text.replace("_", " ")
    text = text.replace("-", " ")
    text = re.sub(r"\s+", " ", text)
    text = text.strip()

    return text


def normalize_symbol(value):
    """
    Normalize PSX stock symbol.
    """

    text = str(value or "").strip().upper()
    text = text.replace(".", "")
    text = text.replace(",", "")
    text = text.replace(":", "")
    text = text.replace(";", "")
    text = text.replace("(", "")
    text = text.replace(")", "")
    text = text.replace("[", "")
    text = text.replace("]", "")

    return text


def clean_share_name(value):
    """
    Clean share/company name.
    """

    text = str(value or "").strip()
    text = re.sub(r"\s+", " ", text)

    return text


def is_valid_symbol(symbol):
    """
    Basic PSX symbol validation.

    This is intentionally stricter for OCR/PDF:
    - 2 to 10 characters
    - uppercase alphanumeric
    - not common English/broker/report word
    - must contain at least one alphabet
    """

    symbol = normalize_symbol(symbol)

    if not symbol:
        return False

    if len(symbol) < 2 or len(symbol) > 10:
        return False

    if not re.match(r"^[A-Z0-9]+$", symbol):
        return False

    if not re.search(r"[A-Z]", symbol):
        return False

    if symbol.lower() in COMMON_WORDS_TO_IGNORE:
        return False

    # Avoid pure dates/years/numbers.
    if re.fullmatch(r"\d+", symbol):
        return False

    return True


def parse_number(value):
    """
    Parse number from text.
    """

    if value is None:
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()

    if not text:
        return 0.0

    text = text.replace(",", "")
    text = text.replace("PKR", "")
    text = text.replace("Rs.", "")
    text = text.replace("Rs", "")
    text = text.replace("rs.", "")
    text = text.replace("rs", "")
    text = text.replace("%", "")
    text = text.strip()

    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]

    match = re.search(r"-?\d+(\.\d+)?", text)

    if not match:
        return 0.0

    try:
        return float(match.group(0))

    except Exception:
        return 0.0


def row_is_empty(row):
    """
    Check whether Excel row is empty.
    """

    for value in row:

        if value is not None and str(value).strip():
            return False

    return True


def get_import_summary(rows):
    """
    Return import preview summary.
    """

    total_rows = len(rows or [])
    total_investment = 0
    total_current = 0
    low_confidence = 0

    for row in rows or []:
        total_investment += parse_number(row.get("investment_value", 0))
        total_current += parse_number(row.get("current_value", 0))

        if int(row.get("confidence", 0)) < 70:
            low_confidence += 1

    return {
        "total_rows": total_rows,
        "total_investment": round(total_investment, 2),
        "total_current": round(total_current, 2),
        "estimated_profit": round(total_current - total_investment, 2),
        "low_confidence_rows": low_confidence,
    }
